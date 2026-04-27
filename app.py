import calendar
import io
import re
from datetime import date

from openpyxl.utils import get_column_letter
import pandas as pd
from fastapi import FastAPI, File, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse

app = FastAPI()

MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}


def parse_period_dates(raw, fallback_year: int) -> tuple[str, str]:
    """Parse Invoice Period into (start_date, end_date) strings dd/mm/yyyy.

    Handles:
      - 'April, 2026'           -> 01/04/2026 - 30/04/2026
      - 'burst April, 2026'     -> 01/04/2026 - 30/04/2026
      - 'January - April, 2026' -> 01/01/2026 - 30/04/2026
      - 'January - April'       -> 01/01/YYYY - 30/04/YYYY (uses fallback_year)
      - empty / unparseable     -> 01/01/YYYY - 31/12/YYYY
    """
    if not isinstance(raw, str) or not raw.strip():
        return f"01/01/{fallback_year}", f"31/12/{fallback_year}"

    raw_clean = raw.strip()

    range_match = re.search(
        r"([A-Za-z]+)\s*[-–]\s*([A-Za-z]+)(?:,?\s*(\d{4}))?", raw_clean
    )
    if range_match:
        m_start = MONTH_MAP.get(range_match.group(1).lower())
        m_end = MONTH_MAP.get(range_match.group(2).lower())
        yr = int(range_match.group(3)) if range_match.group(3) else fallback_year
        if m_start and m_end:
            last_day = calendar.monthrange(yr, m_end)[1]
            return f"01/{m_start:02d}/{yr}", f"{last_day:02d}/{m_end:02d}/{yr}"

    single_match = re.search(r"([A-Za-z]+),?\s*(\d{4})", raw_clean)
    if single_match:
        month_name = single_match.group(1).lower()
        yr = int(single_match.group(2))
        m = MONTH_MAP.get(month_name)
        if m:
            last_day = calendar.monthrange(yr, m)[1]
            return f"01/{m:02d}/{yr}", f"{last_day:02d}/{m:02d}/{yr}"

    return f"01/01/{fallback_year}", f"31/12/{fallback_year}"


def make_row(elem_facturar_id, efc, year, month, start_date, end_date, tipo_cargo, importe):
    return {
        "Elemento a facturar ID": elem_facturar_id,
        "EFC Number": efc,
        "Pendiente de revision Local": "Revisado",
        "Estado de factura": "Provisionado",
        "Año de facturación": year,
        "Mes de facturación": month,
        "Inicio Período de Facturación": start_date,
        "Fin Período de Facturación": end_date,
        "Tipo de Cargo": tipo_cargo,
        "Importe en Curso": float(round(importe, 2)),
    }


def process_excel(contents: bytes) -> bytes:
    xls = pd.ExcelFile(io.BytesIO(contents))

    lookup_prov = pd.read_excel(xls, sheet_name="Informe mes actual Provisiones-")
    efc_to_elem_prov = dict(zip(lookup_prov["EFC Number"], lookup_prov["Elemento a Facturar ID"]))

    efc_to_elem_ext = {}
    if "Informe mes actual EXTORNOS" in xls.sheet_names:
        lookup_ext = pd.read_excel(xls, sheet_name="Informe mes actual EXTORNOS")
        efc_to_elem_ext = dict(zip(lookup_ext["EFC Number"], lookup_ext["Elemento a Facturar ID"]))

    today = date.today()
    current_year = today.year
    current_month = today.month

    provisiones = []
    extornos = []

    for sheet_name in ["Lease", "O&M", "IP"]:
        if sheet_name not in xls.sheet_names:
            continue
        df = pd.read_excel(xls, sheet_name=sheet_name, header=1)

        is_om = sheet_name == "O&M"
        has_nrc = "NRC" in df.columns

        for _, row in df.iterrows():
            ep = row.get("Elemento a Provisionar (EP)")
            ep_ext = row.get("EP EXTORNO  >>> SI APLICA")

            if pd.notna(ep) and isinstance(ep, str) and ep.startswith("EFC"):
                elem_id = efc_to_elem_prov.get(ep)
                if elem_id is None:
                    continue

                inv_period = row.get("Invoice Period")
                start_str, end_str = parse_period_dates(inv_period, current_year)

                mrc_col = "MRC " if "MRC " in df.columns else "MRC"
                mrc_val = row.get(mrc_col, 0)
                if pd.isna(mrc_val):
                    mrc_val = 0
                try:
                    mrc_val = float(mrc_val)
                except (ValueError, TypeError):
                    mrc_val = 0

                if mrc_val > 0:
                    tipo = "O&M" if is_om else "MRC"
                    provisiones.append(make_row(elem_id, ep, current_year, current_month, start_str, end_str, tipo, mrc_val))

                if has_nrc:
                    nrc_val = row.get("NRC", 0)
                    if pd.isna(nrc_val):
                        nrc_val = 0
                    try:
                        nrc_val = float(nrc_val)
                    except (ValueError, TypeError):
                        nrc_val = 0
                    if nrc_val > 0:
                        provisiones.append(make_row(elem_id, ep, current_year, current_month, start_str, end_str, "NRC", nrc_val))

            if pd.notna(ep_ext) and isinstance(ep_ext, str) and ep_ext.startswith("EFC"):
                elem_id = efc_to_elem_ext.get(ep_ext)
                if elem_id is None:
                    continue

                inv_period_ant = row.get("Invoice Period.1")
                start_ant, end_ant = parse_period_dates(inv_period_ant, current_year)

                mrc1_col = "MRC .1" if "MRC .1" in df.columns else "MRC.1"
                mrc1_val = row.get(mrc1_col, 0)
                if pd.isna(mrc1_val):
                    mrc1_val = 0
                try:
                    mrc1_val = float(mrc1_val)
                except (ValueError, TypeError):
                    mrc1_val = 0

                if mrc1_val > 0:
                    tipo = "O&M" if is_om else "MRC"
                    extornos.append(make_row(elem_id, ep_ext, current_year, current_month, start_ant, end_ant, tipo, -mrc1_val))

                if "NRC.1" in df.columns:
                    nrc1_val = row.get("NRC.1", 0)
                    if pd.isna(nrc1_val):
                        nrc1_val = 0
                    try:
                        nrc1_val = float(nrc1_val)
                    except (ValueError, TypeError):
                        nrc1_val = 0
                    if nrc1_val > 0:
                        extornos.append(make_row(elem_id, ep_ext, current_year, current_month, start_ant, end_ant, "NRC", -nrc1_val))

    df_prov = pd.DataFrame(provisiones)
    df_ext = pd.DataFrame(extornos)
    if not df_prov.empty:
        df_prov["Importe en Curso"] = df_prov["Importe en Curso"].astype(float)
    if not df_ext.empty:
        df_ext["Importe en Curso"] = df_ext["Importe en Curso"].astype(float)

    empty_cols = ["Elemento a facturar ID", "EFC Number", "Pendiente de revision Local",
                   "Estado de factura", "Año de facturación", "Mes de facturación",
                   "Inicio Período de Facturación", "Fin Período de Facturación",
                   "Tipo de Cargo", "Importe en Curso"]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        (df_prov if not df_prov.empty else pd.DataFrame(columns=empty_cols)).to_excel(
            writer, sheet_name="Provisiones Positivas", index=False)
        (df_ext if not df_ext.empty else pd.DataFrame(columns=empty_cols)).to_excel(
            writer, sheet_name="Extornos", index=False)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            importe_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == "Importe en Curso":
                    importe_col = col_idx
                    break
            if importe_col:
                for row in ws.iter_rows(min_row=2, min_col=importe_col, max_col=importe_col):
                    for cell in row:
                        if cell.value is not None:
                            cell.value = float(cell.value)
                            cell.number_format = '#,##0.00'

    output.seek(0)
    return output.getvalue()


@app.get("/", response_class=HTMLResponse)
async def index():
    return open("static/index.html", encoding="utf-8").read()


@app.post("/process")
async def process(file: UploadFile = File(...)):
    contents = await file.read()
    try:
        result = process_excel(contents)
    except Exception as e:
        return {"error": str(e)}

    return StreamingResponse(
        io.BytesIO(result),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=carga_salesforce_output.xlsx"},
    )
